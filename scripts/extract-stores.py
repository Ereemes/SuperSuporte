import json
import sys
import openpyxl

xlsx_path = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\sergio.santos\Downloads\Lojas_Organizado_v5 (3).xlsx"
output_path = sys.argv[2] if len(sys.argv) > 2 else "scripts/stores.json"

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb["CADASTRO LOJAS"]

# Row 1 = section headers (merged), Row 2 = column names, Row 3+ = data
headers_row = 2
col_map = {}
for col_idx, cell in enumerate(ws[headers_row], 1):
    if cell.value:
        col_map[str(cell.value).strip()] = col_idx

# Map Excel columns to Store model fields
field_mapping = {
    "Código Loja": "codigo",
    "Nome Fantasia": "nome",
    "Razão Social": "razaoSocial",
    "CNPJ": "cnpj",
    "Status": "status",
    "Unidade de Negócio": "unidadeNegocio",
    "Grupo Financeiro": "grupoFinanceiro",
    "Logradouro": "logradouro",
    "Número": "numero",
    "Complemento (Endereço Lojas)": "complemento",
    "Bairro": "bairro",
    "Cidade": "cidade",
    "UF": "uf",
    "CEP": "cep",
    "Regional": "regional",
    "Diretor": "diretor",
    "Telefone": "telefone",
    "WhatsApp": "whatsapp",
    "E-mail Loja": "emailLoja",
    "E-mail Gerente": "emailGerente",
    "Hora Abertura": "horaAbertura",
    "Hora Fechamento": "horaFechamento",
    "IP": "ip",
    "Link de Internet": "linkInternet",
    "Operadora": "operadora",
    "Tipo Conexão": "tipoConexao",
    "Qtde PDVs": "qtdePdvs",
    "Servidor Local": "servidorLocal",
    "Modelo Equipamento": "modeloEquipamento",
    "Versão MegaStore": "versaoMegastore",
    "Versão Retaguarda": "versaoRetaguarda",
    "Versão Frente": "versaoFrente",
    "Ambiente": "ambiente",
    "Homologação": "homologacao",
    "Observações": "observacoes",
    "Última Atualização": "ultimaAtualizacao",
}

# Build resolved mapping: field_name -> col_index
resolved = {}
for excel_col, model_field in field_mapping.items():
    if excel_col in col_map:
        resolved[model_field] = col_map[excel_col]
    else:
        # Try partial match
        for header, idx in col_map.items():
            if excel_col.lower() in header.lower() or header.lower() in excel_col.lower():
                resolved[model_field] = idx
                break

print(f"Resolved {len(resolved)}/{len(field_mapping)} columns")
missing = set(field_mapping.values()) - set(resolved.keys())
if missing:
    print(f"Missing columns: {missing}")

stores = []
for row_idx in range(headers_row + 1, ws.max_row + 1):
    codigo_col = resolved.get("codigo")
    if not codigo_col:
        continue
    codigo_val = ws.cell(row=row_idx, column=codigo_col).value
    if codigo_val is None or str(codigo_val).strip() == "":
        continue

    store = {}
    for field, col_idx in resolved.items():
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is None:
            store[field] = None
        elif field == "qtdePdvs":
            try:
                store[field] = int(val)
            except (ValueError, TypeError):
                store[field] = None
        elif field == "codigo":
            store[field] = str(int(val)) if isinstance(val, (float, int)) else str(val).strip()
        elif field == "numero":
            if isinstance(val, float):
                store[field] = str(int(val))
            else:
                store[field] = str(val).strip() if val else None
        else:
            store[field] = str(val).strip() if val else None

    if store.get("codigo"):
        stores.append(store)

print(f"Extracted {len(stores)} stores")

with open(output_path, "w", encoding="utf-8") as f:
    json.dump(stores, f, ensure_ascii=False, indent=2)

print(f"Saved to {output_path}")
